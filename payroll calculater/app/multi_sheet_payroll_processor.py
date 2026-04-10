import pandas as pd
import os
import math
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import argparse

# دعم اختياري لتحويل Excel إلى PDF (ويندوز فقط)
try:
    import win32com.client  # type: ignore
    HAS_WIN32 = True
except Exception:
    HAS_WIN32 = False


def process_multi_sheet_payroll(input_file, output_excel_file, output_pdf_file=None, company_name="الإدارة العامة", department="قسم الحسابات"):
    """
    معالجة ملف الرواتب الرئيسي وإنشاء ملف Excel مع ورقة منفصلة لكل موظف
    """
    
    # قراءة ملف Excel الرئيسي باستخدام openpyxl للدقة في القراءة
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(input_file, data_only=True)
        sheet = workbook.active
        print(f"تم قراءة الملف بنجاح. عدد الصفوف: {sheet.max_row}")
    except FileNotFoundError:
        print(f"خطأ: الملف '{input_file}' غير موجود.")
        return
    except Exception as e:
        print(f"خطأ في قراءة الملف: {e}")
        return

    # إنشاء ملف Excel جديد
    wb = Workbook()
    
    # حذف الورقة الافتراضية
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # معالجة كل موظف (بداية من الصف 2 لتخطي العناوين)
    created_sheets = 0
    failed_employees = []
    
    for row_index in range(3, sheet.max_row + 1):  # بداية من الصف 3 لتخطي العناوين
        try:
            # تخطي الصفوف الفارغة
            if all(sheet.cell(row=row_index, column=col).value in [None, ""] for col in range(1, sheet.max_column + 1)):
                continue

            # استخراج البيانات حسب الأعمدة الصحيحة
            employee_code = sheet.cell(row=row_index, column=1).value  # العمود A
            employee_name = sheet.cell(row=row_index, column=2).value  # العمود B
            job_title = sheet.cell(row=row_index, column=3).value      # العمود C
            # قراءة قسم الموظف من العمود C - إذا كان فارغ استخدم القسم الافتراضي
            employee_department = job_title if job_title and str(job_title).strip() != "" else company_name
            
            # تخطي الصفوف التي تحتوي على العناوين
            if str(employee_name).strip() in ['اسم الموظف', 'Name', 'Employee Name', '']:
                continue
            
            # دالة لتحويل القيم إلى أعداد صحيحة (بدون فواصل عشرية)
            def safe_int(value):
                try:
                    if value is None or value == "":
                        return 0
                    return int(float(value))
                except (ValueError, TypeError):
                    return 0
            
            # تحويل جميع البيانات الرقمية إلى أعداد صحيحة
            basic_salary = safe_int(sheet.cell(row=row_index, column=4).value)  # العمود D
            
            # الإجازات
            total_vacation = safe_int(sheet.cell(row=row_index, column=7).value)  # العمود G
            used_vacation = safe_int(sheet.cell(row=row_index, column=8).value)   # العمود H
            remaining_vacation = safe_int(sheet.cell(row=row_index, column=9).value)  # العمود I
            
            # الخصومات
            deduction_hour = safe_int(sheet.cell(row=row_index, column=10).value)   # العمود J
            deduction_day = safe_int(sheet.cell(row=row_index, column=11).value)    # العمود K
            absence_deduction = safe_int(sheet.cell(row=row_index, column=12).value)  # العمود L
            admin_deduction = safe_int(sheet.cell(row=row_index, column=13).value)    # العمود M
            advance_deduction = safe_int(sheet.cell(row=row_index, column=14).value)  # العمود N
            total_deductions = safe_int(sheet.cell(row=row_index, column=15).value)   # العمود O
            
            # الإضافات
            additional_hours = safe_int(sheet.cell(row=row_index, column=16).value)   # العمود P
            additional_days = safe_int(sheet.cell(row=row_index, column=17).value)    # العمود Q
            bonuses = safe_int(sheet.cell(row=row_index, column=18).value)            # العمود R
            transportation = safe_int(sheet.cell(row=row_index, column=19).value)     # العمود S
            punctuality_bonus = safe_int(sheet.cell(row=row_index, column=20).value)  # العمود T
            total_additions = safe_int(sheet.cell(row=row_index, column=21).value)    # العمود U
            net_salary = safe_int(sheet.cell(row=row_index, column=22).value)         # العمود V

            # التأكد من وجود بيانات أساسية
            if not employee_name or employee_name == "":
                continue
                
            # إنشاء الورقة الأولى للموظف (كشف الراتب)
            sheet_name = f"Emp_{employee_code}_Page1"
            ws1 = wb.create_sheet(title=sheet_name)
            ws1.sheet_view.rightToLeft = True
            
            # تنسيق الصفحة الأولى مع البيانات الصحيحة
            create_payroll_sheet(ws1, employee_code, employee_name, job_title, 
                                basic_salary, total_vacation, used_vacation, remaining_vacation,
                                deduction_hour, deduction_day, absence_deduction, admin_deduction, advance_deduction, total_deductions,
                                additional_hours, additional_days, bonuses, transportation, punctuality_bonus, total_additions, 
                                net_salary, company_name, department, notes="لا يوجد", month="أكتوبر", year="2025")
            
            # إنشاء الورقة الثانية للموظف (الصفحة الفارغة مع معلومات BY:RABIE)
            sheet_name2 = f"Emp_{employee_code}_Page2"
            ws2 = wb.create_sheet(title=sheet_name2)
            ws2.sheet_view.rightToLeft = True
            
            # إنشاء الصفحة الثانية الفارغة مع قسم الموظف الحقيقي
            create_second_page(ws2, employee_code, employee_name, employee_department, "أكتوبر", "2025")
            
            created_sheets += 2  # صفحتان لكل موظف
            print(f"✓ تم إنشاء ورقة للموظف: {employee_name} ({employee_code})")
            
        except Exception as e:
            emp_name = f"الصف {row_index}"
            try:
                emp_name = sheet.cell(row=row_index, column=2).value or f"الصف {row_index}"
            except:
                pass
            failed_employees.append((emp_name, str(e)))
            print(f"✗ فشل في معالجة الموظف في {emp_name}: {e}")
    
    # حفظ ملف Excel
    try:
        wb.save(output_excel_file)
        print(f"\n✓ تم حفظ ملف Excel: {output_excel_file}")
        print(f"عدد الأوراق المنشأة: {created_sheets}")
        
        if failed_employees:
            print(f"عدد الموظفين الذين فشلت معالجتهم: {len(failed_employees)}")
            for emp_name, error in failed_employees:
                print(f"  - {emp_name}: {error}")
    except Exception as e:
        print(f"خطأ في حفظ ملف Excel: {e}")
        return
    
    # تحويل إلى PDF إذا طُلب ذلك
    if output_pdf_file:
        convert_excel_to_pdf(output_excel_file, output_pdf_file)


def process_multi_sheet_payroll_with_gui(input_file, output_excel_file, output_pdf_file=None, 
                                        company_name="الإدارة العامة", department="قسم الحسابات", 
                                        notes="لا يوجد", month="أكتوبر", year="2025",
                                        progress_callback=None, status_callback=None):
    """
    دالة معالجة كشوف الرواتب مع دعم الواجهة الرسومية
    """
    
    def update_progress(value):
        if progress_callback:
            progress_callback(value)
            
    def update_status(message):
        if status_callback:
            status_callback(message)
    
    try:
        # قراءة ملف Excel الرئيسي
        update_status("جاري قراءة الملف...")
        update_progress(5)
        
        from openpyxl import load_workbook
        workbook = load_workbook(input_file, data_only=True)
        sheet = workbook.active
        update_progress(10)
        
        # إنشاء ملف Excel جديد
        update_status("إنشاء ملف Excel جديد...")
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        update_progress(15)
        
        # حساب عدد الموظفين لتحديث التقدم
        total_rows = 0
        for row_index in range(3, sheet.max_row + 1):
            if not all(sheet.cell(row=row_index, column=col).value in [None, ""] for col in range(1, sheet.max_column + 1)):
                employee_name = sheet.cell(row=row_index, column=2).value
                if employee_name and str(employee_name).strip() not in ['اسم الموظف', 'Name', 'Employee Name', '']:
                    total_rows += 1
        
        update_status(f"تم العثور على {total_rows} موظف")
        update_progress(20)
        
        # معالجة كل موظف
        created_sheets = 0
        failed_employees = []
        
        for row_index in range(3, sheet.max_row + 1):
            try:
                # تخطي الصفوف الفارغة
                if all(sheet.cell(row=row_index, column=col).value in [None, ""] for col in range(1, sheet.max_column + 1)):
                    continue

                # استخراج البيانات
                employee_code = sheet.cell(row=row_index, column=1).value
                employee_name = sheet.cell(row=row_index, column=2).value
                job_title = sheet.cell(row=row_index, column=3).value
                # قراءة قسم الموظف من العمود C (العمود الثالث) - إذا كان فارغ استخدم القسم الافتراضي
                employee_department = job_title if job_title and str(job_title).strip() != "" else department
                
                # تخطي الصفوف التي تحتوي على العناوين
                if str(employee_name).strip() in ['اسم الموظف', 'Name', 'Employee Name', '']:
                    continue
                
                # دالة لتحويل القيم إلى أعداد صحيحة مع التعامل مع المسافات الفارغة
                def safe_int(value):
                    try:
                        if value is None or value == "":
                            return 0
                        # التعامل مع المسافات الفارغة والنصوص المحتوية على مسافات فقط
                        if isinstance(value, str) and value.strip() == "":
                            return 0
                        return int(float(value))
                    except (ValueError, TypeError):
                        return 0
                
                # تحويل جميع البيانات الرقمية إلى أعداد صحيحة
                basic_salary = safe_int(sheet.cell(row=row_index, column=4).value)
                total_vacation = safe_int(sheet.cell(row=row_index, column=7).value)
                used_vacation = safe_int(sheet.cell(row=row_index, column=8).value)
                remaining_vacation = safe_int(sheet.cell(row=row_index, column=9).value)
                deduction_hour = safe_int(sheet.cell(row=row_index, column=10).value)
                deduction_day = safe_int(sheet.cell(row=row_index, column=11).value)
                absence_deduction = safe_int(sheet.cell(row=row_index, column=12).value)
                admin_deduction = safe_int(sheet.cell(row=row_index, column=13).value)
                advance_deduction = safe_int(sheet.cell(row=row_index, column=14).value)
                total_deductions = safe_int(sheet.cell(row=row_index, column=15).value)
                additional_hours = safe_int(sheet.cell(row=row_index, column=16).value)
                additional_days = safe_int(sheet.cell(row=row_index, column=17).value)
                bonuses = safe_int(sheet.cell(row=row_index, column=18).value)
                transportation = safe_int(sheet.cell(row=row_index, column=19).value)
                punctuality_bonus = safe_int(sheet.cell(row=row_index, column=20).value)
                total_additions = safe_int(sheet.cell(row=row_index, column=21).value)
                net_salary = safe_int(sheet.cell(row=row_index, column=22).value)

                # التأكد من وجود بيانات أساسية
                if not employee_name or employee_name == "":
                    continue
                    
                # تحديث التقدم
                progress_percent = 20 + (created_sheets / total_rows) * 60
                update_progress(progress_percent)
                update_status(f"معالجة الموظف: {employee_name}")
                
                # إنشاء الورقة الأولى للموظف (كشف الراتب)
                sheet_name = f"Emp_{employee_code}_Page1"
                ws1 = wb.create_sheet(title=sheet_name)
                ws1.sheet_view.rightToLeft = True
                
                # تنسيق الصفحة الأولى مع الملاحظات المخصصة
                create_payroll_sheet(ws1, employee_code, employee_name, job_title, 
                                    basic_salary, total_vacation, used_vacation, remaining_vacation,
                                    deduction_hour, deduction_day, absence_deduction, admin_deduction, advance_deduction, total_deductions,
                                    additional_hours, additional_days, bonuses, transportation, punctuality_bonus, total_additions, 
                                    net_salary, company_name, department, notes, month, year)
                
                # إنشاء الورقة الثانية للموظف (الصفحة الفارغة مع معلومات BY:RABIE)
                sheet_name2 = f"Emp_{employee_code}_Page2"
                ws2 = wb.create_sheet(title=sheet_name2)
                ws2.sheet_view.rightToLeft = True
                
                # إنشاء الصفحة الثانية الفارغة مع قسم الموظف الحقيقي
                create_second_page(ws2, employee_code, employee_name, employee_department, month, year)
                
                created_sheets += 2  # عدد الأوراق الآن مضاعف (صفحتان لكل موظف)
                
            except Exception as e:
                emp_name = f"الصف {row_index}"
                try:
                    emp_name = sheet.cell(row=row_index, column=2).value or f"الصف {row_index}"
                except:
                    pass
                failed_employees.append((emp_name, str(e)))
        
        # حفظ ملف Excel
        update_status("حفظ ملف Excel...")
        update_progress(85)
        wb.save(output_excel_file)
        
        # تحويل إلى PDF
        if output_pdf_file:
            update_status("تحويل إلى PDF...")
            update_progress(90)
            convert_excel_to_pdf(output_excel_file, output_pdf_file)
        
        update_progress(100)
        return True
        
    except Exception as e:
        update_status(f"خطأ: {str(e)}")
        return False


def create_payroll_sheet(ws, employee_code, employee_name, job_title, 
                        basic_salary, total_vacation, used_vacation, remaining_vacation,
                        deduction_hour, deduction_day, absence_deduction, admin_deduction, advance_deduction, total_deductions,
                        additional_hours, additional_days, bonuses, transportation, punctuality_bonus, total_additions,
                        net_salary, company_name, department, notes="لا يوجد", month="أكتوبر", year="2025"):
    
    # استخدام الشهر والعام المرسلين من الواجهة
    month_text = f"عن شهر {month} لعام {year}"
    """إنشاء تصميم كشف الراتب لموظف واحد مطابق للنموذج المرفق مع البيانات الفعلية"""
    
    import math
    
    # تنسيقات الخطوط والألوان (مطابقة لملف test.py)
    header_font_red = Font(bold=True, size=14, color="990000")  # أحمر داكن
    bold_font = Font(bold=True, size=12, color="000000")
    normal_font = Font(size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    
    # الحدود
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # الألوان (مطابقة لملف test.py)
    header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")     # أرجواني فاتح
    earnings_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")   # أزرق سماوي
    deductions_fill = PatternFill(start_color="FFF0F5", end_color="FFF0F5", fill_type="solid") # وردي فاتح
    label_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")       # أزرق فاتح للعناوين
    net_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")         # أخضر فاتح للصافي
    zebra_fill_1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")     # أبيض
    zebra_fill_2 = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")     # رمادي فاتح
    
    def apply_cell_style(cell, font=None, alignment=None, border=None, fill=None):
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if fill:
            cell.fill = fill

    # عرض الأعمدة وارتفاع الصفوف
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 30

    for row in range(1, 17):
        ws.row_dimensions[row].height = 18
    ws.row_dimensions[17].height = 20
    ws.row_dimensions[18].height = 20

    # العناوين
    ws.merge_cells("B1:E1")
    ws["B1"] = "الإدارة العامة"
    apply_cell_style(ws["B1"], font=header_font_red, alignment=center_align, fill=header_fill)

    ws.merge_cells("B2:E2")
    ws["B2"] = "قسم الحسابات"
    apply_cell_style(ws["B2"], font=header_font_red, alignment=center_align, fill=header_fill)

    ws.merge_cells("B3:E3")
    ws["B3"] = "مُفردات الراتب الشهري"
    apply_cell_style(ws["B3"], font=header_font_red, alignment=center_align, fill=header_fill)
    
    # الصف 4: دمج من B4 إلى E4 (صف فارغ للتنسيق)
    ws.merge_cells("B4:E4")
    ws["B4"] = ""  # صف فارغ مدموج

    # بيانات الموظف
    ws["B5"] = "كود الموظف:"
    ws["C5"] = employee_code
    apply_cell_style(ws["B5"], font=bold_font, fill=label_fill, border=thin_border, alignment=right_align)
    apply_cell_style(ws["C5"], font=normal_font, border=thin_border, alignment=right_align)

    ws["D5"] = "اسم الموظف:"
    ws["E5"] = employee_name
    apply_cell_style(ws["D5"], font=bold_font, fill=label_fill, border=thin_border, alignment=right_align)
    apply_cell_style(ws["E5"], font=normal_font, border=thin_border, alignment=right_align)

    ws["B6"] = "القسم:"
    ws["C6"] = job_title
    apply_cell_style(ws["B6"], font=bold_font, fill=label_fill, border=thin_border, alignment=right_align)
    apply_cell_style(ws["C6"], font=normal_font, border=thin_border, alignment=right_align)

    ws.merge_cells("D6:E6")
    ws["D6"] = month_text  # استخدام النص المحسوب تلقائياً
    apply_cell_style(ws["D6"], font=bold_font, alignment=center_align, fill=header_fill)

    # ملاحظات (مخصصة)
    ws.merge_cells("B7:E7")
    ws["B7"] = f"ملاحظات: {notes}"
    apply_cell_style(ws["B7"], font=Font(size=10, italic=True), alignment=center_align,
                     fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"))

    # الاستحقاقات (البيانات الفعلية)
    ws.merge_cells("B8:C8")
    ws["B8"] = "الاستحقاقات"
    apply_cell_style(ws["B8"], font=bold_font, alignment=center_align, fill=earnings_fill, border=thin_border)

    earnings_data = [
        ("الراتب الأساسي", basic_salary),
        ("الإضافي الساعات", additional_hours),
        ("الإضافي الأيام", additional_days),
        ("المكافآت", bonuses),
        ("بدل مواصلات", transportation),
        ("حافز انتظام", punctuality_bonus),
        ("إجمالي الإضافات", total_additions),
    ]

    for idx, (label, value) in enumerate(earnings_data, start=9):
        cell_b = ws.cell(row=idx, column=2, value=label)
        cell_c = ws.cell(row=idx, column=3, value=value)
        fill_style = zebra_fill_1 if idx % 2 == 1 else zebra_fill_2
        apply_cell_style(cell_b, font=bold_font, fill=label_fill, border=thin_border, alignment=right_align)
        apply_cell_style(cell_c, font=normal_font, fill=fill_style, border=thin_border, alignment=right_align)

    # الاستقطاعات (البيانات الفعلية)
    ws.merge_cells("D8:E8")
    ws["D8"] = "الاستقطاعات"
    apply_cell_style(ws["D8"], font=bold_font, alignment=center_align, fill=deductions_fill, border=thin_border)

    deductions_data = [
        ("خصم بالساعة", deduction_hour),
        ("خصم باليوم", deduction_day),
        ("خصم غياب", absence_deduction),
        ("خصم إداري", admin_deduction),
        ("السلف", advance_deduction),
        ("إجمالي الخصومات", total_deductions),
        ("", ""),  # صف فارغ
    ]

    for idx, (label, value) in enumerate(deductions_data, start=9):
        if label:  # تخطي الصفوف الفارغة
            cell_d = ws.cell(row=idx, column=4, value=label)
            cell_e = ws.cell(row=idx, column=5, value=value)
            fill_style = zebra_fill_1 if idx % 2 == 1 else zebra_fill_2
            apply_cell_style(cell_d, font=bold_font, fill=label_fill, border=thin_border, alignment=right_align)
            apply_cell_style(cell_e, font=normal_font, fill=fill_style, border=thin_border, alignment=right_align)

    # صافي الراتب كعدد صحيح (بدون فواصل عشرية)
    try:
        net_value = int(float(net_salary)) if net_salary else 0
    except (ValueError, TypeError):
        net_value = 0
    english_digits = str(net_value).translate(str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789"))

    # صافي الراتب (النص بالأسود الغامق)
    ws.merge_cells("B16:E16")
    ws["B16"] = f"صافي الراتب المستحق: {english_digits} جنيه"
    apply_cell_style(ws["B16"], 
                     font=Font(bold=True, size=14, color="000000"),  # أسود غامق
                     alignment=center_align,
                     fill=net_fill, 
                     border=thin_border)

    # التواقيع
    for col in ["B", "C", "D", "E"]:
        ws.merge_cells(f"{col}17:{col}18")
    ws["B17"] = "توقيع الموظف"
    ws["C17"] = "المُعد"
    ws["D17"] = "الصندوق"
    ws["E17"] = "الإدارة المالية"

    for col in ["B", "C", "D", "E"]:
        apply_cell_style(ws[f"{col}17"], font=bold_font, alignment=center_align, fill=label_fill, border=thin_border)
        apply_cell_style(ws[f"{col}18"], font=bold_font, alignment=center_align, border=thin_border)

    # تطبيق الحد العريض على الجدول كله (B1 إلى E18)
    for row in range(1, 19):
        for col in range(2, 6):  # B=2, C=3, D=4, E=5
            cell = ws.cell(row=row, column=col)
            border = Border(
                left=Side(style='medium') if col == 2 else Side(style='thin'),
                right=Side(style='medium') if col == 5 else Side(style='thin'),
                top=Side(style='medium') if row == 1 else Side(style='thin'),
                bottom=Side(style='medium') if row == 18 else Side(style='thin')
            )
            cell.border = border


def create_second_page(ws, employee_code, employee_name, department, month, year):
    """إنشاء الصفحة الثانية الفارغة مع معلومات الموظف في الأسفل"""
    
    # تنسيقات الخط
    bold_font = Font(bold=True, size=12, color="000000")
    normal_font = Font(size=11, color="000000")
    large_font = Font(bold=True, size=16, color="990000")
    
    # محاذاة
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # دالة مساعدة لتطبيق التنسيقات
    def apply_style(cell, font=None, alignment=None):
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
    
    # ترك معظم الصفحة فارغة والبدء من الصف 35 (أسفل الصفحة)
    
    # معلومات الموظف في أسفل الصفحة
    ws["B35"] = f"اسم الموظف: {employee_name}"
    ws["B36"] = f"كود الموظف: {employee_code}"
    ws["B37"] = f"القسم: {department}"
    ws["B38"] = f"التاريخ: {month} {year}"
    
    # تطبيق التنسيقات على معلومات الموظف
    for row in range(35, 39):
        apply_style(ws[f"B{row}"], font=normal_font, alignment=right_align)
    
    # دمج خلايا لكلمة BY:RABIE لتكون أكثر وضوحاً
    ws.merge_cells("D40:E40")
    ws["D40"] = "BY: RABIE"
    apply_style(ws["D40"], font=large_font, alignment=center_align)


def validate_payroll_data(input_file, progress_callback=None, status_callback=None):
    """
    التحقق من صحة بيانات الرواتب والعمليات الحسابية
    إرجاع: (نجح الفحص، قائمة الأخطاء، تقرير مفصل)
    """
    
    def update_progress(value):
        if progress_callback:
            progress_callback(value)
            
    def update_status(message):
        if status_callback:
            status_callback(message)
    
    try:
        update_status("🔍 جاري فحص ملف البيانات...")
        update_progress(5)
        
        # قراءة الملف
        from openpyxl import load_workbook
        workbook = load_workbook(input_file, data_only=True)
        sheet = workbook.active
        update_progress(15)
        
        errors = []
        warnings = []
        total_employees = 0
        valid_employees = 0
        
        update_status("🧮 فحص العمليات الحسابية...")
        
        # فحص كل موظف
        for row_index in range(3, sheet.max_row + 1):
            try:
                # تخطي الصفوف الفارغة
                if all(sheet.cell(row=row_index, column=col).value in [None, ""] for col in range(1, sheet.max_column + 1)):
                    continue

                # استخراج البيانات الأساسية
                employee_code = sheet.cell(row=row_index, column=1).value
                employee_name = sheet.cell(row=row_index, column=2).value
                
                # تخطي صفوف العناوين
                if str(employee_name).strip() in ['اسم الموظف', 'Name', 'Employee Name', '']:
                    continue
                
                if not employee_name or employee_name == "":
                    continue
                    
                total_employees += 1
                
                # تحديث التقدم
                progress_percent = 15 + (total_employees * 70 / 250)  # افتراض 250 موظف كحد أقصى
                update_progress(progress_percent)
                update_status(f"🔍 فحص الموظف: {employee_name}")
                
                # دالة لتحويل القيم مع التعامل مع الأخطاء والمسافات الفارغة
                def safe_float(value, field_name=""):
                    try:
                        if value is None or value == "":
                            return 0.0
                        # التعامل مع المسافات الفارغة والنصوص المحتوية على مسافات فقط
                        if isinstance(value, str) and value.strip() == "":
                            return 0.0
                        return float(value)
                    except (ValueError, TypeError):
                        errors.append(f"الصف {row_index} - {employee_name}: قيمة غير صحيحة في {field_name}: '{value}'")
                        return 0.0                # استخراج جميع القيم المالية
                basic_salary = safe_float(sheet.cell(row=row_index, column=4).value, "الراتب الأساسي")
                total_vacation = safe_float(sheet.cell(row=row_index, column=7).value, "إجمالي الإجازات")
                used_vacation = safe_float(sheet.cell(row=row_index, column=8).value, "الإجازات المستخدمة")
                remaining_vacation = safe_float(sheet.cell(row=row_index, column=9).value, "الإجازات المتبقية")
                
                # الخصومات
                deduction_hour = safe_float(sheet.cell(row=row_index, column=10).value, "خصم ساعات")
                deduction_day = safe_float(sheet.cell(row=row_index, column=11).value, "خصم أيام")
                absence_deduction = safe_float(sheet.cell(row=row_index, column=12).value, "خصم غياب")
                admin_deduction = safe_float(sheet.cell(row=row_index, column=13).value, "خصم إداري")
                advance_deduction = safe_float(sheet.cell(row=row_index, column=14).value, "السلف")
                total_deductions = safe_float(sheet.cell(row=row_index, column=15).value, "إجمالي الخصومات")
                
                # الإضافات
                additional_hours = safe_float(sheet.cell(row=row_index, column=16).value, "إضافي ساعات")
                additional_days = safe_float(sheet.cell(row=row_index, column=17).value, "إضافي أيام")
                bonuses = safe_float(sheet.cell(row=row_index, column=18).value, "المكافآت")
                transportation = safe_float(sheet.cell(row=row_index, column=19).value, "بدل مواصلات")
                punctuality_bonus = safe_float(sheet.cell(row=row_index, column=20).value, "حافز انتظام")
                total_additions = safe_float(sheet.cell(row=row_index, column=21).value, "إجمالي الإضافات")
                net_salary = safe_float(sheet.cell(row=row_index, column=22).value, "صافي الراتب")

                # التحققات الحسابية مع الحسابات الصحيحة
                employee_has_errors = False
                
                # حساب قيمة اليوم والساعة من الراتب الأساسي
                daily_wage = basic_salary / 26  # الراتب الأساسي ÷ 26 يوم
                hourly_wage = daily_wage / 8    # اليوم ÷ 8 ساعات
                
                # حد التسامح للفروقات البسيطة (5 جنيهات أو أقل)
                TOLERANCE = 5.0
                
                # 1. فحص إجمالي الخصومات (مع ضرب الساعات والأيام بالمعاملات)
                calculated_hour_deduction = deduction_hour * hourly_wage  # عدد الساعات × أجر الساعة
                calculated_day_deduction = deduction_day * daily_wage     # عدد الأيام × أجر اليوم  
                calculated_absence_deduction = absence_deduction * daily_wage  # أيام الغياب × أجر اليوم
                
                calculated_deductions = calculated_hour_deduction + calculated_day_deduction + calculated_absence_deduction + admin_deduction + advance_deduction
                deduction_difference = abs(calculated_deductions - total_deductions)
                
                if deduction_difference > TOLERANCE:  # تسامح مع الفروقات أقل من 5 جنيهات
                    errors.append(f"الصف {row_index} - {employee_name}: خطأ في إجمالي الخصومات - المسجل: {total_deductions:.2f}, المحسوب: {calculated_deductions:.2f}, الفرق: {deduction_difference:.2f} (ساعات: {calculated_hour_deduction:.2f} + أيام: {calculated_day_deduction:.2f} + غياب: {calculated_absence_deduction:.2f} + إداري: {admin_deduction:.2f} + سلف: {advance_deduction:.2f})")
                    employee_has_errors = True
                elif deduction_difference > 0.01:  # إذا كان الفرق بسيط، أضفه للتحذيرات
                    warnings.append(f"الصف {row_index} - {employee_name}: فرق بسيط في الخصومات ({deduction_difference:.2f} جنيه) - مقبول")
                
                # 2. فحص إجمالي الإضافات (مع ضرب الساعات والأيام بالمعاملات)
                calculated_hour_addition = additional_hours * hourly_wage  # ساعات إضافية × أجر الساعة
                calculated_day_addition = additional_days * daily_wage     # أيام إضافية × أجر اليوم
                
                calculated_additions = calculated_hour_addition + calculated_day_addition + bonuses + transportation + punctuality_bonus
                addition_difference = abs(calculated_additions - total_additions)
                
                if addition_difference > TOLERANCE:  # تسامح مع الفروقات أقل من 5 جنيهات
                    errors.append(f"الصف {row_index} - {employee_name}: خطأ في إجمالي الإضافات - المسجل: {total_additions:.2f}, المحسوب: {calculated_additions:.2f}, الفرق: {addition_difference:.2f} (ساعات: {calculated_hour_addition:.2f} + أيام: {calculated_day_addition:.2f} + مكافآت: {bonuses:.2f} + مواصلات: {transportation:.2f} + انتظام: {punctuality_bonus:.2f})")
                    employee_has_errors = True
                elif addition_difference > 0.01:  # إذا كان الفرق بسيط، أضفه للتحذيرات
                    warnings.append(f"الصف {row_index} - {employee_name}: فرق بسيط في الإضافات ({addition_difference:.2f} جنيه) - مقبول")
                
                # 3. فحص صافي الراتب
                calculated_net = basic_salary + total_additions - total_deductions
                net_difference = abs(calculated_net - net_salary)
                
                if net_difference > TOLERANCE:  # تسامح مع الفروقات أقل من 5 جنيهات
                    errors.append(f"الصف {row_index} - {employee_name}: خطأ في صافي الراتب - المسجل: {net_salary:.2f}, المحسوب: {calculated_net:.2f}, الفرق: {net_difference:.2f} (أساسي: {basic_salary:.2f} + إضافات: {total_additions:.2f} - خصومات: {total_deductions:.2f})")
                    employee_has_errors = True
                elif net_difference > 0.01:  # إذا كان الفرق بسيط، أضفه للتحذيرات
                    warnings.append(f"الصف {row_index} - {employee_name}: فرق بسيط في صافي الراتب ({net_difference:.2f} جنيه) - مقبول")
                
                # 4. فحص الإجازات
                if abs((total_vacation - used_vacation) - remaining_vacation) > 0.01:
                    warnings.append(f"الصف {row_index} - {employee_name}: تحذير في حساب الإجازات المتبقية")
                
                # 5. فحص القيم المنطقية
                if basic_salary < 0:
                    errors.append(f"الصف {row_index} - {employee_name}: الراتب الأساسي لا يمكن أن يكون سالباً")
                    employee_has_errors = True
                
                if net_salary < 0:
                    warnings.append(f"الصف {row_index} - {employee_name}: تحذير - صافي الراتب سالب")
                
                # 6. فحص كود الموظف
                if not employee_code:
                    errors.append(f"الصف {row_index} - {employee_name}: كود الموظف مفقود")
                    employee_has_errors = True
                
                if not employee_has_errors:
                    valid_employees += 1
                    
            except Exception as e:
                errors.append(f"الصف {row_index}: خطأ في معالجة البيانات - {str(e)}")
        
        # فحص التكرار في أكواد الموظفين
        update_status("🔍 فحص تكرار أكواد الموظفين...")
        update_progress(90)
        
        employee_codes = []
        for row_index in range(3, sheet.max_row + 1):
            employee_code = sheet.cell(row=row_index, column=1).value
            employee_name = sheet.cell(row=row_index, column=2).value
            
            if employee_code and employee_name and str(employee_name).strip() not in ['اسم الموظف', 'Name', 'Employee Name', '']:
                if employee_code in employee_codes:
                    errors.append(f"كود موظف مكرر: {employee_code} (الصف {row_index})")
                else:
                    employee_codes.append(employee_code)
        
        update_progress(100)
        
        # إنشاء التقرير
        report = f"""
📊 تقرير فحص بيانات الرواتب
═════════════════════════════════

📈 الإحصائيات العامة:
• إجمالي الموظفين المفحوصين: {total_employees}
• الموظفين الصحيحين: {valid_employees}
• الموظفين بأخطاء: {total_employees - valid_employees}

💡 ملاحظة: الفروقات أقل من 5 جنيهات تُعتبر مقبولة (تقريب طبيعي)

{'✅ لا توجد أخطاء - جميع البيانات صحيحة!' if not errors else '❌ تم اكتشاف أخطاء:'}
"""
        
        if errors:
            report += f"\n🚨 الأخطاء المكتشفة ({len(errors)}):\n"
            report += "─" * 40 + "\n"
            for i, error in enumerate(errors[:10], 1):  # أول 10 أخطاء فقط
                report += f"{i}. {error}\n"
            if len(errors) > 10:
                report += f"... و {len(errors) - 10} أخطاء أخرى\n"
        
        if warnings:
            report += f"\n⚠️ التحذيرات ({len(warnings)}):\n"
            report += "─" * 40 + "\n"
            for i, warning in enumerate(warnings[:5], 1):  # أول 5 تحذيرات
                report += f"{i}. {warning}\n"
            if len(warnings) > 5:
                report += f"... و {len(warnings) - 5} تحذيرات أخرى\n"
        
        # تحديث الحالة النهائية
        if errors:
            update_status(f"❌ تم اكتشاف {len(errors)} أخطاء")
        else:
            update_status("✅ جميع البيانات صحيحة!")
        
        success = len(errors) == 0
        return success, errors, warnings, report
        
    except Exception as e:
        error_msg = f"خطأ في فحص الملف: {str(e)}"
        update_status(error_msg)
        return False, [error_msg], [], f"❌ فشل في فحص الملف: {str(e)}"


def convert_excel_to_pdf(excel_file, pdf_file):
    """تحويل ملف Excel إلى PDF"""
    if not HAS_WIN32:
        print("تحويل PDF يتطلب pywin32 وMicrosoft Excel على Windows")
        return False
    
    try:
        print(f"جاري تحويل {excel_file} إلى PDF...")
        
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # فتح الملف
        wb = excel.Workbooks.Open(os.path.abspath(excel_file))
        
        # تصدير كـ PDF
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf_file))  # 0 = xlTypePDF
        
        # إغلاق الملف والتطبيق
        wb.Close(SaveChanges=False)
        excel.Quit()
        
        print(f"✓ تم تحويل الملف إلى PDF: {pdf_file}")
        return True
        
    except Exception as e:
        print(f"خطأ في تحويل PDF: {e}")
        return False


def main():
    parser = argparse.ArgumentParser(description="معالج كشوف الرواتب متعدد الأوراق")
    parser.add_argument("-i", "--input", dest="input_file", 
                       default="input-master.xlsx", 
                       help="ملف Excel المدخل (افتراضي: input-master.xlsx)")
    parser.add_argument("-o", "--output", dest="output_excel", 
                       default="output-master-processed.xlsx", 
                       help="ملف Excel المخرج (افتراضي: output-master-processed.xlsx)")
    parser.add_argument("-p", "--pdf", dest="output_pdf", 
                       help="ملف PDF المخرج (اختياري)")
    parser.add_argument("--company", dest="company_name", 
                       default="الإدارة العامة", 
                       help="اسم الشركة/المؤسسة")
    parser.add_argument("--department", dest="department", 
                       default="قسم الحسابات", 
                       help="اسم القسم")
    
    args = parser.parse_args()
    
    # التحقق من وجود الملف المدخل
    if not os.path.exists(args.input_file):
        print(f"خطأ: الملف المدخل '{args.input_file}' غير موجود.")
        return
    
    # معالجة الرواتب
    process_multi_sheet_payroll(
        input_file=args.input_file,
        output_excel_file=args.output_excel,
        output_pdf_file=args.output_pdf,
        company_name=args.company_name,
        department=args.department
    )


if __name__ == "__main__":
    main()